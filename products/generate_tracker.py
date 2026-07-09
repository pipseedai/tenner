#!/usr/bin/env python3
"""Tenner product 1: UK Sole Trader Finance Tracker (.xlsx).

Generates two editions into products/dist/ (gitignored):
  - tenner-uk-sole-trader-tracker-FULL.xlsx   (paid)
  - tenner-uk-sole-trader-tracker-LITE.xlsx   (free sample)

All tax figures verified against gov.uk on 2026-07-09:
  - Allowable expense categories: gov.uk/expenses-if-youre-self-employed
  - Mileage flat rates 2026-27: 55p first 10,000 miles, 25p after;
    motorcycles 24p: gov.uk/simpler-income-tax-simplified-expenses/vehicles
  - Working from home flat rates: £10 (25-50h), £18 (51-100h), £26 (101h+):
    gov.uk/simpler-income-tax-simplified-expenses/working-from-home
  - Trading allowance £1,000: gov.uk/expenses-if-youre-self-employed
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

INK = "1F2430"          # near-black text
BRAND = "D96C2C"        # warm orange (a crisp tenner is orange-brown)
BRAND_DARK = "A84E17"
PAPER = "FDF8F2"        # warm paper
BAND = "F4E3D3"         # subheader band
OK = "2E7D4F"
GBP = '£#,##0.00;[Red]-£#,##0.00'

ROWS = 200  # data rows per log sheet

CATEGORIES = [
    "Office costs",
    "Travel costs",
    "Clothing expenses",
    "Staff costs",
    "Reselling goods",
    "Financial costs",
    "Business premises costs",
    "Marketing",
    "Training courses",
    "Other (check gov.uk)",
]

CATEGORY_NOTES = {
    "Office costs": "Stationery, phone bills, software, postage.",
    "Travel costs": "Fuel, parking, train or bus fares (NOT commuting; use Mileage sheet if claiming flat-rate mileage instead of actual vehicle costs).",
    "Clothing expenses": "Uniforms and protective clothing only — not everyday clothes.",
    "Staff costs": "Salaries or subcontractor costs.",
    "Reselling goods": "Stock or raw materials you buy to sell on.",
    "Financial costs": "Insurance, bank charges, accountancy.",
    "Business premises costs": "Heating, lighting, business rates. Working from home? Use the Home Office sheet flat rates instead, or apportion actual costs.",
    "Marketing": "Advertising, website costs, free samples.",
    "Training courses": "Courses related to your existing business, e.g. refresher courses.",
    "Other (check gov.uk)": "Only if genuinely wholly and exclusively for business — check gov.uk/expenses-if-youre-self-employed.",
}

thin = Side(style="thin", color="D9CDBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_title(ws, title, subtitle, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BRAND)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(size=9, italic=True, color=BRAND_DARK)
    s.fill = PatternFill("solid", fgColor=BAND)
    s.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[2].height = 18
    ws.sheet_view.showGridLines = False


def header_row(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=INK)
        c.alignment = Alignment(vertical="center", indent=1)
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 20
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def data_region(ws, first_row, n_rows, n_cols, money_cols=(), date_cols=()):
    for r in range(first_row, first_row + n_rows):
        for col in range(1, n_cols + 1):
            c = ws.cell(row=r, column=col)
            c.border = border
            c.font = Font(size=10)
            if r % 2 == 0:
                c.fill = PatternFill("solid", fgColor=PAPER)
            if col in money_cols:
                c.number_format = GBP
            if col in date_cols:
                c.number_format = "DD/MM/YYYY"


def total_cell(ws, row, col, label_col, label, formula):
    lc = ws.cell(row=row, column=label_col, value=label)
    lc.font = Font(bold=True, size=10)
    tc = ws.cell(row=row, column=col, value=formula)
    tc.font = Font(bold=True, size=11, color=BRAND_DARK)
    tc.number_format = GBP
    tc.fill = PatternFill("solid", fgColor=BAND)
    tc.border = border


def sheet_income(wb):
    ws = wb.create_sheet("Income")
    style_title(ws, "Income", "Log every payment you receive. Money in only — expenses have their own sheet.", 5)
    header_row(ws, 4, ["Date", "Description", "Client / source", "Amount", "Notes"], [12, 38, 22, 14, 30])
    data_region(ws, 5, ROWS, 5, money_cols=(4,), date_cols=(1,))
    total_cell(ws, 5 + ROWS + 1, 4, 2, "Total income", f"=SUM(D5:D{4 + ROWS})")
    dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0",
                        errorTitle="Positive amounts", error="Income entries should be positive.")
    ws.add_data_validation(dv)
    dv.add(f"D5:D{4 + ROWS}")
    return ws


def sheet_expenses(wb):
    ws = wb.create_sheet("Expenses")
    style_title(ws, "Expenses", "HMRC allowable-expense categories (dropdown). Only costs wholly and exclusively for business.", 6)
    header_row(ws, 4, ["Date", "Description", "Category", "Amount", "Business %", "Notes"], [12, 34, 26, 14, 12, 28])
    data_region(ws, 5, ROWS, 6, money_cols=(4,), date_cols=(1,))
    cat_list = '"' + ",".join(CATEGORIES) + '"'
    dv = DataValidation(type="list", formula1=cat_list, allow_blank=True,
                        errorTitle="Pick a category", error="Choose one of the HMRC categories from the list.")
    ws.add_data_validation(dv)
    dv.add(f"C5:C{4 + ROWS}")
    for r in range(5, 5 + ROWS):
        ws.cell(row=r, column=5).number_format = "0%"
    total_cell(ws, 5 + ROWS + 1, 4, 2, "Total expenses", f"=SUM(D5:D{4 + ROWS})")
    return ws


def sheet_mileage(wb):
    ws = wb.create_sheet("Mileage")
    style_title(ws, "Mileage (simplified expenses)",
                "2026-27 flat rates, verified gov.uk: cars/goods 55p first 10,000 business miles, then 25p. Motorcycles 24p all miles. Use EITHER flat-rate mileage OR actual vehicle costs — not both.", 5)
    header_row(ws, 4, ["Date", "Journey (from → to)", "Purpose", "Vehicle", "Miles"], [12, 36, 24, 14, 10])
    data_region(ws, 5, ROWS, 5, date_cols=(1,))
    dv = DataValidation(type="list", formula1='"Car/van,Motorcycle"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"D5:D{4 + ROWS}")
    last = 4 + ROWS
    r = last + 2
    labels_formulas = [
        ("Car/van miles", f'=SUMIFS(E5:E{last},D5:D{last},"Car/van")', "0"),
        ("Motorcycle miles", f'=SUMIFS(E5:E{last},D5:D{last},"Motorcycle")', "0"),
        ("Car/van claim (55p ≤10k, 25p after)", f"=MIN(B{r},10000)*0.55+MAX(B{r}-10000,0)*0.25", GBP),
        ("Motorcycle claim (24p)", f"=B{r + 1}*0.24", GBP),
        ("Total mileage claim", f"=B{r + 2}+B{r + 3}", GBP),
    ]
    for i, (label, formula, fmt) in enumerate(labels_formulas):
        ws.cell(row=r + i, column=1, value=label).font = Font(bold=(i == 4), size=10)
        c = ws.cell(row=r + i, column=2, value=formula)
        c.number_format = fmt
        c.font = Font(bold=True, size=10, color=BRAND_DARK if i == 4 else INK)
        if i == 4:
            c.fill = PatternFill("solid", fgColor=BAND)
    return ws


def sheet_home_office(wb):
    ws = wb.create_sheet("Home Office")
    style_title(ws, "Working from home (simplified expenses)",
                "gov.uk flat rates: 25-50 hrs/month £10 · 51-100 hrs £18 · 101+ hrs £26. Needs 25+ hrs/month. Phone & internet are claimed separately on the Expenses sheet.", 4)
    header_row(ws, 4, ["Month", "Hours worked from home", "Flat rate", "Note"], [16, 22, 12, 40])
    months = ["April", "May", "June", "July", "August", "September", "October",
              "November", "December", "January", "February", "March"]
    for i, m in enumerate(months):
        r = 5 + i
        ws.cell(row=r, column=1, value=m)
        f = (f'=IF(B{r}="","",IF(B{r}<25,0,IF(B{r}<=50,10,IF(B{r}<=100,18,26))))')
        c = ws.cell(row=r, column=3, value=f)
        c.number_format = GBP
        ws.cell(row=r, column=4, value="Below 25 hrs = £0 under simplified expenses" if i == 0 else None)
    data_region(ws, 5, 12, 4, money_cols=(3,))
    total_cell(ws, 18, 3, 1, "Total home-office claim", "=SUM(C5:C16)")
    return ws


def sheet_dashboard(wb, lite=False):
    ws = wb.create_sheet("Dashboard", 1)
    style_title(ws, "Year position", "Live totals from your logs. Estimated figures — not tax advice.", 4)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 60
    last = 4 + ROWS
    rows = [("Total income", f"=SUM(Income!D5:D{last})")]
    rows.append(("Total expenses", f"=SUM(Expenses!D5:D{last})"))
    if not lite:
        rows.append(("Mileage claim", f"=Mileage!B{last + 6}"))
        rows.append(("Home-office claim", "='Home Office'!C18"))
        rows.append(("Estimated profit before tax",
                     "=B5-B6-B7-B8"))
    else:
        rows.append(("Estimated profit before tax", "=B5-B6"))
    r = 5
    for label, formula in rows:
        ws.cell(row=r, column=1, value=label).font = Font(size=11, bold=("profit" in label.lower()))
        c = ws.cell(row=r, column=2, value=formula)
        c.number_format = GBP
        c.font = Font(bold=True, size=12, color=BRAND_DARK if "profit" in label.lower() else INK)
        c.border = border
        if "profit" in label.lower():
            c.fill = PatternFill("solid", fgColor=BAND)
        r += 1
    r += 1
    tip = ws.cell(row=r, column=1,
                  value="Trading allowance: if your total self-employment income is under £1,000 you may not need to report it — but you can't claim expenses AND the allowance. Check gov.uk.")
    tip.font = Font(size=9, italic=True, color=BRAND_DARK)
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=2)
    tip.alignment = Alignment(wrap_text=True, vertical="top")
    if not lite:
        ws.cell(row=4, column=4, value="Expenses by category").font = Font(bold=True, size=11)
        for i, cat in enumerate(CATEGORIES):
            rr = 5 + i
            ws.cell(row=rr, column=4, value=cat).font = Font(size=10)
            c = ws.cell(row=rr, column=5, value=f'=SUMIFS(Expenses!D5:D{last},Expenses!C5:C{last},D{rr})')
            c.number_format = GBP
            c.font = Font(size=10)
            ws.column_dimensions["E"].width = 14
    return ws


def sheet_start(wb, lite=False):
    ws = wb.create_sheet("Start Here", 0)
    style_title(ws, "UK Sole Trader Finance Tracker" + (" — free sample" if lite else ""),
                "Built by Tenner, an AI agent growing £10 into a business · ko-fi.com/crisptenner", 2)
    ws.column_dimensions["A"].width = 110
    lines = [
        "",
        "WHAT THIS IS",
        "A clean place to keep the money side of self-employment: income, expenses in HMRC's own categories,"
        + ("" if lite else " flat-rate mileage, working-from-home claims,") + " and a live year position.",
        "",
        "HOW TO USE IT",
        "1. Log every payment in on the Income sheet.",
        "2. Log every business cost on the Expenses sheet — pick the category from the dropdown.",
    ]
    if not lite:
        lines += [
            "3. Business journeys go in Mileage (rates: 55p/mile first 10,000 business miles for cars & vans, 25p after; motorcycles 24p — 2026-27 rates, verified on gov.uk).",
            "4. Work from home 25+ hours a month? Enter monthly hours in Home Office for the gov.uk flat rate.",
            "5. Dashboard shows your live position. That's it. No macros, works in Excel, Google Sheets and LibreOffice.",
        ]
    else:
        lines += [
            "3. Dashboard shows your live position. No macros; works in Excel, Google Sheets and LibreOffice.",
            "",
            "THIS IS THE FREE SAMPLE. The full version adds: flat-rate Mileage log (2026-27 rates), Home Office claims, a category-by-category dashboard, and the expense category guide.",
            "Get it at ko-fi.com/crisptenner",
        ]
    lines += [
        "",
        "THE SMALL PRINT",
        "This is a record-keeping template, not tax advice, and it doesn't know your situation. Figures verified against gov.uk guidance on 9 July 2026; rules change — check gov.uk or ask a qualified accountant.",
        "Sources: gov.uk/expenses-if-youre-self-employed · gov.uk/simpler-income-tax-simplified-expenses",
        "",
        "WHO MADE THIS",
        "I'm Tenner — an AI agent handed £10 and told to grow it. This tracker is how I'd keep my own books if I were you. Questions or a custom version: crisptenner@fastmail.com",
    ]
    r = 4
    for line in lines:
        c = ws.cell(row=r, column=1, value=line if line else None)
        if line.isupper() and line:
            c.font = Font(bold=True, size=11, color=BRAND_DARK)
        else:
            c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = None
        r += 1
    return ws


def sheet_category_guide(wb):
    ws = wb.create_sheet("Category guide")
    style_title(ws, "HMRC expense categories, in plain English",
                "Straight from gov.uk/expenses-if-youre-self-employed (checked 9 July 2026).", 2)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 90
    header_row(ws, 4, ["Category", "What goes in it"], [26, 90])
    for i, cat in enumerate(CATEGORIES):
        r = 5 + i
        ws.cell(row=r, column=1, value=cat).font = Font(bold=True, size=10)
        c = ws.cell(row=r, column=2, value=CATEGORY_NOTES[cat])
        c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=1).border = border
        c.border = border
    return ws


def build(path, lite=False):
    wb = Workbook()
    wb.remove(wb.active)
    sheet_start(wb, lite=lite)
    sheet_dashboard(wb, lite=lite)
    sheet_income(wb)
    sheet_expenses(wb)
    if not lite:
        sheet_mileage(wb)
        sheet_home_office(wb)
        sheet_category_guide(wb)
    wb.save(path)
    print("wrote", path)


if __name__ == "__main__":
    import pathlib
    dist = pathlib.Path(__file__).parent / "dist"
    dist.mkdir(exist_ok=True)
    build(dist / "tenner-uk-sole-trader-tracker-FULL.xlsx", lite=False)
    build(dist / "tenner-uk-sole-trader-tracker-LITE.xlsx", lite=True)
